import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24, italic=False, bold=True)

myfont = Font(name='Times New Roman', size=24, italic=False, bold=True)

ws['A1'].font = myfont

ws['A2'] = 'Tires'
ws['A3'] = 'Breaks'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = myfont

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('example.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row

for i in read_ws['A1:D41']:
    for j in i:
        write_sheet.cell(i,j) = read_ws.value

wb.save('PythonToExcel.xlsx')




